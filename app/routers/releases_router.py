from typing import Annotated
import openpyxl
from openpyxl.styles import Border, Side, Font, PatternFill

from fastapi import APIRouter, Depends, HTTPException

from sqlalchemy.orm import Session
from starlette.responses import FileResponse

from sql_app import releases_service
from sql_app.channels_service import get_channel
from sql_app.database import get_database

from sql_app.models.user import RolesEnum
from sql_app.platforms_service import get_platform
from sql_app.releases_service import get_all_releases, update_release, get_release, get_all_release_types
from schemas import ReleaseStageCreate, User, ReleaseStageOut, ReleaseTypeOut, ReleaseStageOutWithFeature, \
    PaginationReleaseStages, ReleaseStatusENUM
from auth import get_current_user
import logg_config

logger = logg_config.get_logger(__name__)
router = APIRouter(prefix="/releases", tags=["releases"])
get_current_user = Annotated[User, Depends(get_current_user)]
db_session = Annotated[Session, Depends(get_database)]


@router.post("/", response_model=ReleaseStageOut)
def create_release(stage: ReleaseStageCreate,
                   current_user: get_current_user,
                   db: db_session,
                   ):
    logger.info("User %s is attempting to create a new release: %s", current_user.username, stage.name)
    if current_user.role not in [RolesEnum.ADMIN.value, RolesEnum.RELEASE_MANAGER.value]:
        logger.warning("User %s does not have enough permissions", current_user.username)
        raise HTTPException(status_code=403, detail="Not enough permissions")
    if get_release(name=stage.name, db=db):
        logger.warning("Release stage with name %s already exists", stage.name)
        raise HTTPException(status_code=400, detail="Release stage with this name already exists")
    if not get_channel(channel_id=stage.channel_id, db=db):
        logger.warning("Channel not found with ID: %d", stage.channel_id)
        raise HTTPException(status_code=404, detail="Channel not found")
    if not get_platform(platform_id=stage.platform_id, db=db):
        logger.warning("Platform not found with ID: %d", stage.platform_id)
        raise HTTPException(status_code=404, detail="Platform not found")
    release = releases_service.create_release(stage=stage, db=db)
    logger.info("Release %s created successfully by user %s", stage.name, current_user.username)
    return release


@router.get("/all", response_model=PaginationReleaseStages, status_code=200)
def get_all_releases(db: db_session,
                     platform_id: int | None = None,
                     channel_id: int | None = None,
                     status: ReleaseStatusENUM | None = None,
                     page: int = 1,
                     page_size: int = 50):
    logger.info("Fetching all releases")
    data, page_size, total = releases_service.get_all_releases(db=db,
                                                               platform_id=platform_id,
                                                               channel_id=channel_id,
                                                               status = status.value if status else None,
                                                               page=page,
                                                               page_size=page_size)
    result = []
    for release in data:
        result_row = ReleaseStageOutWithFeature.from_orm(release.Release)
        result_row.features = release.features
        result.append(result_row)
    return {'data': result, 'page': page, 'page_size': page_size, 'total': total}


@router.get('/{release_id}', status_code=200)
def get_release(release_id: int, db: db_session):
    logger.info("Getting release with ID: %d", release_id)
    release = releases_service.get_release_with_features(release_id=release_id, db=db)
    if not release:
        logger.warning("Release not found with ID: %d", release_id)
        raise HTTPException(status_code=404, detail="Release not found")
    response = ReleaseStageOutWithFeature.from_orm(release.Release)
    response.features = release.features
    return response


@router.delete("/{release_id}", status_code=204)
def delete_release(release_id: int,
                   current_user: get_current_user,
                   db: db_session):
    logger.info("User %s is attempting to delete release with ID: %d", current_user.username, release_id)
    if current_user.role not in [RolesEnum.ADMIN.value, RolesEnum.RELEASE_MANAGER.value]:
        logger.warning("User %s does not have enough permissions", current_user.username)
        raise HTTPException(status_code=403, detail="Not enough permissions")
    release = releases_service.get_release(release_id=release_id, db=db)
    if not release:
        logger.warning("Release not found with ID: %d", release_id)
        raise HTTPException(status_code=404, detail="Release not found")
    releases_service.delete_release(stage_id=release_id, db=db)
    logger.info("Release with ID: %d deleted successfully by user %s", release_id, current_user.username)
    return None


@router.get("/{release_id}/report", response_class=FileResponse)
def generate_report(release_id: int, db: db_session):
    release = get_release(release_id=release_id, db=db)
    if not release:
        logger.warning("Release not found with ID: %d", release_id)
        raise HTTPException(status_code=404, detail="Release not found")

    features = releases_service.get_release_with_features(release_id=release_id, db=db).features

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = release.name

    # Feature details
    feature_headers = ["Имя фичи", "Ключ фичи", "Статус", "Тип фичи"]
    for col_num, header in enumerate(feature_headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, size=14)
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    thin_border = Side(border_style='thin', color='000000')
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    for row_num, feature in enumerate(features, 2):
        for col_num in range(1, 5):
            cell = ws.cell(row=row_num, column=col_num)
            if col_num == 1:
                cell.value = feature['name']
            elif col_num == 2:
                cell.value = feature['jira_key']
            elif col_num == 3:
                cell.value = feature['status']
            elif col_num == 4:
                cell.value = feature['feature_type_id']

    ws.auto_filter.ref = ws.dimensions
    ws.auto_filter.add_filter_column(1, [feature['status'] for feature in features])

    for row in ws['A1:D6']:
        for cell in row:
            cell.border = border

    file_path = "release_report.xlsx"
    wb.save(file_path)
    logger.info("Report generated successfully")
    return FileResponse(file_path,
                        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        filename=file_path)


@router.put("/{stage_id}")
def update_release(stage_id: int,
                   name: str | None,
                   description: str | None,
                   start_date: str | None,
                   end_date: str | None,
                   current_user: get_current_user,
                   db: db_session):
    logger.info("User %s is attempting to update release with ID: %d", current_user.username, stage_id)
    if current_user.role not in [RolesEnum.ADMIN.value, RolesEnum.RELEASE_MANAGER.value]:
        logger.warning("User %s does not have enough permissions", current_user.username)
        raise HTTPException(status_code=403, detail="Not enough permissions")
    updated_release = update_release(stage_id=stage_id,
                                     name=name,
                                     description=description,
                                     start_date=start_date,
                                     end_date=end_date,
                                     db=db)
    logger.info("Release with ID: %d updated successfully by user %s", stage_id, current_user.username)
    return updated_release


@router.get('/types/', response_model=list[ReleaseTypeOut], status_code=200)
def get_release_types(db: db_session):
    logger.info("Fetching all release types")
    release_types = get_all_release_types(db=db)
    return release_types


@router.post("/types/", response_model=ReleaseTypeOut, status_code=201)
def create_release_type(name: str,
                        platform_id: int,
                        channel_id: int,
                        current_user: get_current_user,
                        db: db_session):
    logger.info("User %s is attempting to create a new release type: %s", current_user.username, name)
    if current_user.role not in [RolesEnum.ADMIN.value, RolesEnum.RELEASE_MANAGER.value]:
        logger.warning("User %s does not have enough permissions", current_user.username)
        raise HTTPException(status_code=403, detail="Not enough permissions")
    if not get_platform(platform_id=platform_id, db=db):
        logger.warning("Platform not found with ID: %d", platform_id)
        raise HTTPException(status_code=404, detail="Platform not found")
    if not get_channel(channel_id=channel_id, db=db):
        logger.warning("Channel not found with ID: %d", channel_id)
        raise HTTPException(status_code=404, detail="Channel not found")
    new_release_type = releases_service.create_release_type(name=name, platform_id=platform_id, channel_id=channel_id,
                                                            db=db)
    logger.info("Release type %s created successfully by user %s", name, current_user.username)
    return new_release_type
